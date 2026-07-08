from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - user-facing guard
    raise SystemExit(
        "Не установлен openpyxl. Выполните: python -m pip install -r requirements.txt"
    ) from exc

from .config import DIRECTION_SHORT_NAMES, DIRECTIONS, QUOTAS, load_direction_short_names
from .models import ParseResult

BASE_HEADERS = [
    "ФИО",
    "Email",
    "Телефон",
    "Сумма баллов",
    "Сумма баллов по предметам",
    "Сумма баллов за инд.дост.",
    "Согласие на зачисление",
    "Примечание",
]
EMAIL_COLUMN = 2
EXTRA_QUOTAS_HEADER = "Дополнительные квоты"
SHORT_NAMES_BY_FACULTY = load_direction_short_names()


@dataclass(frozen=True)
class ExistingWorkbookData:
    path: Path
    faculty: str
    notes_by_applicant: dict[str, str]
    priorities_by_applicant: dict[str, dict[str, str]]


@dataclass(frozen=True)
class PriorityCheckResult:
    changed: int
    unchanged: int
    added: int
    missing: int


def write_result_xlsx(
    parse_result: ParseResult,
    output_path: str | Path,
    notes_by_applicant: dict[str, str] | None = None,
) -> Path:
    path = Path(output_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    readme_sheet = workbook.active
    readme_sheet.title = "ReadMe"
    applicants_sheet = workbook.create_sheet("Абитуриенты")
    statistics_sheet = workbook.create_sheet("Статистика")

    _fill_readme(readme_sheet, parse_result)
    _fill_applicants(applicants_sheet, parse_result, notes_by_applicant or {})
    _fill_statistics(statistics_sheet, parse_result)

    workbook.save(path)
    return path


def read_existing_workbook(workbook_path: str | Path) -> ExistingWorkbookData:
    path = Path(workbook_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Ожидается существующий XLSX-файл.")

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if "Абитуриенты" not in workbook.sheetnames:
            raise ValueError("В XLSX-файле отсутствует лист 'Абитуриенты'.")

        sheet = workbook["Абитуриенты"]
        headers = {
            str(cell.value).strip(): cell.column
            for cell in sheet[1]
            if cell.value is not None and str(cell.value).strip()
        }
        required_headers = {"ФИО", "Email", EXTRA_QUOTAS_HEADER}
        missing_headers = required_headers - headers.keys()
        if missing_headers:
            missing = ", ".join(sorted(missing_headers))
            raise ValueError(f"В листе 'Абитуриенты' отсутствуют столбцы: {missing}.")

        notes: dict[str, str] = {}
        priorities: dict[str, dict[str, str]] = {}
        priority_columns = {
            header: column
            for header, column in headers.items()
            if column >= headers[EXTRA_QUOTAS_HEADER]
        }
        note_column = headers.get("Примечание")

        for row_index in range(2, sheet.max_row + 1):
            fio = _cell_text(sheet.cell(row=row_index, column=headers["ФИО"]).value)
            email = _cell_text(sheet.cell(row=row_index, column=headers["Email"]).value)
            applicant_key = _applicant_key(email, fio)
            if not applicant_key:
                continue

            note = (
                _cell_text(sheet.cell(row=row_index, column=note_column).value)
                if note_column is not None
                else ""
            )
            if note:
                notes[applicant_key] = _merge_note_values(notes.get(applicant_key, ""), note)

            priorities[applicant_key] = {
                header: value
                for header, column in priority_columns.items()
                if (value := _cell_text(sheet.cell(row=row_index, column=column).value))
            }

        return ExistingWorkbookData(
            path=path,
            faculty=_read_workbook_faculty(workbook),
            notes_by_applicant=notes,
            priorities_by_applicant=priorities,
        )
    finally:
        workbook.close()


def compare_priorities(
    previous: ExistingWorkbookData,
    current: ExistingWorkbookData,
) -> PriorityCheckResult:
    previous_keys = set(previous.priorities_by_applicant)
    current_keys = set(current.priorities_by_applicant)
    common_keys = previous_keys & current_keys
    changed = sum(
        previous.priorities_by_applicant[key] != current.priorities_by_applicant[key]
        for key in common_keys
    )
    return PriorityCheckResult(
        changed=changed,
        unchanged=len(common_keys) - changed,
        added=len(current_keys - previous_keys),
        missing=len(previous_keys - current_keys),
    )


def _fill_readme(sheet, parse_result: ParseResult) -> None:
    rows = [
        ("Что это за файл", "Выгрузка абитуриентов из PDF 'Список деканам'. Каждый абитуриент показан один раз: повторные заявления объединяются по Email."),
        ("Лист Абитуриенты", "Основной рабочий лист для обзвона, фильтрации и сортировки поступающих."),
        ("ФИО", "ФИО абитуриента из исходного PDF."),
        ("Email", "Используется как уникальный идентификатор абитуриента. По нему объединяются повторяющиеся заявления."),
        ("Телефон", "Телефоны приведены к единому виду. Если телефонов несколько, каждый указан с новой строки внутри ячейки."),
        ("Баллы", "Сумма баллов, сумма баллов по предметам и баллы за индивидуальные достижения вынесены в отдельные узкие столбцы."),
        ("Согласие на зачисление", "Значение 'Да' означает, что в исходном PDF было отмечено согласие на зачисление."),
        ("Примечание", "Свободное поле для рабочих заметок. При обновлении существующего XLSX примечания переносятся по Email."),
        (
            "Обновление файла",
            "Если при запуске указать существующий XLSX, приложение обновит список, приоритеты и статистику по новому PDF, сохранив примечания.",
        ),
        ("Дополнительные квоты", "Не-бюджетные заявления записаны кратко: сокращение направления, дефис, приоритет и код квоты. Например: ИВТ-7П."),
        ("Бюджетные направления", "Столбцы с сокращениями направлений показывают только бюджетный приоритет числом."),
        (
            "Цветная заливка",
            "В бюджетных столбцах условное форматирование подсвечивает приоритеты 1-3: 1 зеленым, 2 светло-зеленым, 3 желтым.",
        ),
        (
            "Как пользоваться",
            "Можно фильтровать по согласию, сортировать по баллам или по приоритету нужного направления, а также смотреть дополнительные квоты отдельным столбцом.",
        ),
        (
            "Сортировка по приоритету",
            "Чтобы увидеть самых заинтересованных абитуриентов по направлению, отсортируйте соответствующий бюджетный столбец по возрастанию: 1, затем 2, затем 3.",
        ),
        (
            "Лист Статистика",
            "На листе собраны сведения о документе, количество заявлений, уникальных абитуриентов и сводки по направлениям/квотам.",
        ),
        (
            "Сокращения направлений",
            "Короткие названия направлений можно редактировать в файле direction_short_names.ini рядом с приложением.",
        ),
    ]

    directions = _directions_for_result(parse_result)

    for row_index, (name, value) in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1, value=name)
        sheet.cell(row=row_index, column=2, value=value)

    sheet.cell(row=len(rows) + 2, column=1, value="Направления")
    for index, direction in enumerate(directions, start=len(rows) + 3):
        sheet.cell(row=index, column=1, value=direction)

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 110
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _fill_applicants(
    sheet,
    parse_result: ParseResult,
    notes_by_applicant: dict[str, str],
) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    alternate_row_fill = PatternFill("solid", fgColor="F4F8FB")
    bold = Font(bold=True)
    directions = _directions_for_result(parse_result)

    headers = BASE_HEADERS + [EXTRA_QUOTAS_HEADER]
    directions = _directions_for_result(parse_result)
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    direction_start_column = len(headers) + 1
    col_index = direction_start_column
    for direction in directions:
        direction_cell = sheet.cell(
            row=1,
            column=col_index,
            value=_direction_short_name(direction, parse_result.faculty),
        )
        direction_cell.font = bold
        direction_cell.fill = header_fill
        direction_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col_index += 1

    for row_index, applicant in enumerate(parse_result.applicants, start=2):
        values = [
            applicant.fio,
            applicant.email,
            applicant.phone,
            applicant.total_score,
            applicant.subject_score,
            applicant.achievement_score,
            "Да" if applicant.consent else "",
            notes_by_applicant.get(_applicant_key(applicant.email, applicant.fio), ""),
            _format_extra_quotas(applicant.priorities, directions, parse_result.faculty),
        ]
        for value_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=value_index, value=value)

        col_index = direction_start_column
        for direction in directions:
            priority_cell = sheet.cell(
                row=row_index,
                column=col_index,
                value=_format_budget_priority(applicant.priorities, direction),
            )
            priority_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            col_index += 1

        if row_index % 2 == 0:
            for cell in sheet[row_index]:
                cell.fill = alternate_row_fill

    sheet.freeze_panes = "D2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    _add_priority_conditional_formatting(sheet, direction_start_column)

    widths = {
        1: 34,
        2: 28,
        3: 25,
        4: 9,
        5: 11,
        6: 11,
        7: 16,
        8: 28,
        9: 30,
    }
    for col_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_index)].width = width
    for col_index in range(direction_start_column, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col_index)].width = 10

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center" if cell.row == 1 or cell.column >= direction_start_column else None,
                vertical="top",
                wrap_text=cell.column != EMAIL_COLUMN,
            )


def _format_extra_quotas(priorities: dict[tuple[str, str], str], directions: list[str], faculty: str) -> str:
    values = []
    for direction in directions:
        direction_short_name = _direction_short_name(direction, faculty)
        for quota in QUOTAS:
            if quota == "Бюджет":
                continue
            priority = priorities.get((direction, quota), "")
            for priority_part in _split_priority_values(priority):
                values.append(f"{direction_short_name}-{priority_part}")
    return "; ".join(values)


def _format_budget_priority(priorities: dict[tuple[str, str], str], direction: str) -> int | str:
    priority = priorities.get((direction, "Бюджет"), "")
    if priority.isdigit():
        return int(priority)
    return priority


def _split_priority_values(priority: str) -> list[str]:
    return [part.strip() for part in priority.split(";") if part.strip()]


def _add_priority_conditional_formatting(sheet, direction_start_column: int) -> None:
    if sheet.max_row < 2:
        return

    priority_fills = {
        1: PatternFill(fill_type="solid", start_color="FF63BE7B", end_color="FF63BE7B"),
        2: PatternFill(fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE"),
        3: PatternFill(fill_type="solid", start_color="FFFFEB9C", end_color="FFFFEB9C"),
    }
    for col_index in range(direction_start_column, sheet.max_column + 1):
        column_letter = get_column_letter(col_index)
        cell_range = f"{column_letter}2:{column_letter}{sheet.max_row}"
        for priority, fill in priority_fills.items():
            sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="equal", formula=[str(priority)], fill=fill),
            )


def _fill_statistics(sheet, parse_result: ParseResult) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="BDD7EE")
    alternate_row_fill = PatternFill("solid", fgColor="F4F8FB")
    bold = Font(bold=True)
    directions = _directions_for_result(parse_result)

    current_row = 1
    current_row = _write_summary_statistics(
        sheet,
        parse_result,
        current_row,
        section_fill,
        header_fill,
        alternate_row_fill,
        bold,
    )
    current_row += 2
    current_row = _write_records_by_direction_and_quota(
        sheet,
        parse_result,
        directions,
        current_row,
        section_fill,
        header_fill,
        alternate_row_fill,
        bold,
        table_name="ApplicationsByDirectionQuota",
    )
    current_row += 2
    current_row = _write_budget_priority_statistics(
        sheet,
        parse_result,
        directions,
        current_row,
        section_fill,
        header_fill,
        alternate_row_fill,
        bold,
        table_name="BudgetPriorityStats",
    )
    current_row += 2
    _write_extra_quota_statistics(
        sheet,
        parse_result,
        directions,
        current_row,
        section_fill,
        header_fill,
        alternate_row_fill,
        bold,
        table_name="ExtraQuotaStats",
    )

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    widths = {
        1: 18,
        2: 18,
        3: 18,
        4: 18,
        5: 18,
        6: 20,
        7: 18,
        8: 18,
    }
    for col_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_index)].width = width

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center" if cell.column > 1 else None,
                vertical="top",
                wrap_text=True,
            )


def _write_summary_statistics(
    sheet,
    parse_result: ParseResult,
    start_row: int,
    section_fill: PatternFill,
    header_fill: PatternFill,
    alternate_row_fill: PatternFill,
    bold: Font,
) -> int:
    rows = [
        ("Факультет", parse_result.faculty),
        ("Дата формирования XLSX", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Страниц PDF", parse_result.pages_count),
        ("Всего заявлений", len(parse_result.records)),
        ("Уникальных абитуриентов", len(parse_result.applicants)),
        ("Заявлений с согласием", sum(1 for record in parse_result.records if record.consent)),
        ("Уникальных абитуриентов с согласием", sum(1 for applicant in parse_result.applicants if applicant.consent)),
        ("Конкурсных групп", len(parse_result.groups)),
    ]
    return _write_table(
        sheet,
        start_row,
        "Итого",
        ["Показатель", "Значение"],
        rows,
        section_fill,
        header_fill,
        alternate_row_fill,
        bold,
    )


def _write_records_by_direction_and_quota(
    sheet,
    parse_result: ParseResult,
    directions: list[str],
    start_row: int,
    section_fill: PatternFill,
    header_fill: PatternFill,
    alternate_row_fill: PatternFill,
    bold: Font,
    table_name: str | None = None,
) -> int:
    counts = Counter((record.direction, record.quota) for record in parse_result.records)
    headers = ["Направление", *QUOTAS, "Итого"]
    rows = []
    for direction in directions:
        values = [counts[(direction, quota)] for quota in QUOTAS]
        rows.append([_direction_short_name(direction, parse_result.faculty), *values, sum(values)])

    quota_totals = [sum(counts[(direction, quota)] for direction in directions) for quota in QUOTAS]
    rows.append(["Итого", *quota_totals, sum(quota_totals)])

    return _write_table(
        sheet,
        start_row,
        "Заявления по направлениям и квотам",
        headers,
        rows,
        section_fill,
        header_fill,
        alternate_row_fill,
        bold,
        table_name=table_name,
    )


def _write_budget_priority_statistics(
    sheet,
    parse_result: ParseResult,
    directions: list[str],
    start_row: int,
    section_fill: PatternFill,
    header_fill: PatternFill,
    alternate_row_fill: PatternFill,
    bold: Font,
    table_name: str | None = None,
) -> int:
    priority_columns = ["1", "2", "3", "4+", "Без приоритета/прочее"]
    counts = {
        direction: Counter(_budget_priority_bucket(applicant.priorities.get((direction, "Бюджет"), "")) for applicant in parse_result.applicants if applicant.priorities.get((direction, "Бюджет"), ""))
        for direction in directions
    }
    headers = ["Направление", *priority_columns, "Итого"]
    rows = []
    for direction in directions:
        values = [counts[direction][column] for column in priority_columns]
        rows.append([_direction_short_name(direction, parse_result.faculty), *values, sum(values)])

    totals = [sum(counts[direction][column] for direction in directions) for column in priority_columns]
    rows.append(["Итого", *totals, sum(totals)])

    return _write_table(
        sheet,
        start_row,
        "Уникальные абитуриенты по бюджетным приоритетам",
        headers,
        rows,
        section_fill,
        header_fill,
        alternate_row_fill,
        bold,
        table_name=table_name,
    )


def _write_extra_quota_statistics(
    sheet,
    parse_result: ParseResult,
    directions: list[str],
    start_row: int,
    section_fill: PatternFill,
    header_fill: PatternFill,
    alternate_row_fill: PatternFill,
    bold: Font,
    table_name: str | None = None,
) -> int:
    extra_quotas = [quota for quota in QUOTAS if quota != "Бюджет"]
    counts = {
        direction: Counter(
            quota
            for applicant in parse_result.applicants
            for quota in extra_quotas
            if applicant.priorities.get((direction, quota), "")
        )
        for direction in directions
    }
    headers = ["Направление", *extra_quotas, "Итого"]
    rows = []
    for direction in directions:
        values = [counts[direction][quota] for quota in extra_quotas]
        rows.append([_direction_short_name(direction, parse_result.faculty), *values, sum(values)])

    totals = [sum(counts[direction][quota] for direction in directions) for quota in extra_quotas]
    rows.append(["Итого", *totals, sum(totals)])

    return _write_table(
        sheet,
        start_row,
        "Дополнительные квоты",
        headers,
        rows,
        section_fill,
        header_fill,
        alternate_row_fill,
        bold,
        table_name=table_name,
    )


def _write_table(
    sheet,
    start_row: int,
    title: str,
    headers: list[str],
    rows: list[tuple | list],
    section_fill: PatternFill,
    header_fill: PatternFill,
    alternate_row_fill: PatternFill,
    bold: Font,
    table_name: str | None = None,
) -> int:
    title_cell = sheet.cell(row=start_row, column=1, value=title)
    title_cell.font = bold
    title_cell.fill = section_fill
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))

    header_row = start_row + 1
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_index, value=header)
        cell.font = bold
        cell.fill = header_fill

    for row_offset, row_values in enumerate(rows, start=1):
        row_index = header_row + row_offset
        for col_index, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            if row_values[0] == "Итого":
                cell.font = bold
        if row_offset % 2 == 0 and row_values[0] != "Итого":
            for cell in sheet[row_index]:
                cell.fill = alternate_row_fill

    end_row = header_row + len(rows)
    if table_name is not None:
        _add_table_filter(sheet, table_name, header_row, end_row, len(headers))

    return end_row


def _read_workbook_faculty(workbook) -> str:
    if "Статистика" not in workbook.sheetnames:
        return ""

    for label, value in workbook["Статистика"].iter_rows(
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if _cell_text(label) == "Факультет":
            return _cell_text(value)
    return ""


def _applicant_key(email: str, fio: str) -> str:
    normalized_email = email.strip().casefold()
    if normalized_email:
        return f"email:{normalized_email}"

    normalized_fio = re.sub(r"\s+", " ", fio).strip().casefold()
    return f"fio:{normalized_fio}" if normalized_fio else ""


def _cell_text(value) -> str:
    return "" if value is None else str(value).strip()


def _merge_note_values(existing: str, new_value: str) -> str:
    if not existing:
        return new_value
    if new_value == existing or new_value in existing.splitlines():
        return existing
    return f"{existing}\n{new_value}"


def _budget_priority_bucket(priority: str) -> str:
    values = _split_priority_values(priority)
    if not values:
        return "Без приоритета/прочее"

    numeric_values = []
    for value in values:
        if value.isdigit():
            numeric_values.append(int(value))

    if not numeric_values:
        return "Без приоритета/прочее"

    best_priority = min(numeric_values)
    if best_priority in {1, 2, 3}:
        return str(best_priority)
    return "4+"


def _directions_for_result(parse_result: ParseResult) -> list[str]:
    directions: list[str] = []
    for record in parse_result.records:
        if record.direction and record.direction not in directions:
            directions.append(record.direction)

    known_order = [direction for direction in DIRECTIONS if direction in directions]
    other_directions = [direction for direction in directions if direction not in known_order]
    return known_order + other_directions


def _direction_short_name(direction: str, faculty: str) -> str:
    faculty_short_names = SHORT_NAMES_BY_FACULTY.get(faculty, {})
    if direction in faculty_short_names:
        return faculty_short_names[direction]
    if direction in DIRECTION_SHORT_NAMES:
        return DIRECTION_SHORT_NAMES[direction]

    cleaned = re.sub(r"\([^)]*\)", "", direction).strip()
    words = re.findall(r"[A-Za-zА-Яа-яЁё]+", cleaned)
    if not words:
        return direction[:12]
    if len(words) == 1:
        return words[0] if len(words[0]) <= 12 else words[0][:12]

    acronym = "".join(word[0].upper() for word in words)
    return acronym[:12]


def _add_table_filter(sheet, table_name: str, header_row: int, end_row: int, columns_count: int) -> None:
    end_column = get_column_letter(columns_count)
    table = Table(displayName=table_name, ref=f"A{header_row}:{end_column}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)
